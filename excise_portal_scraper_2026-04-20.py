"""
Excise Portal Scraper — Full PAD Flow Conversion
Fully automated: login → navigate panels → filter → download → combine.

Requirements:
    pip install playwright openpyxl
"""

import os
import sys
import time
import glob
import shutil
import subprocess
import threading
import queue

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
except ImportError:
    print("ERROR: tkinter is not available. Reinstall Python and check 'tcl/tk' option.")
    input("Press Enter to close...")
    sys.exit(1)

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    tk.messagebox.showerror("Missing Package", "playwright is not installed.\n\nRun this in Command Prompt:\n  pip install playwright") if 'tk' in dir() else None
    print("\nERROR: playwright is not installed.")
    print("Fix: open Command Prompt and run:")
    print("  pip install playwright")
    input("\nPress Enter to close...")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    # Patch openpyxl's _convert to tolerate unconvertible style IDs in SAP-exported xlsx files
    try:
        import openpyxl.descriptors.base as _oxdb
        _orig_convert = _oxdb._convert
        def _tolerant_convert(expected_type, value):
            try:
                return _orig_convert(expected_type, value)
            except TypeError:
                if expected_type is int:
                    try:
                        return int(float(str(value)))
                    except (ValueError, TypeError):
                        return 0
                raise
        _oxdb._convert = _tolerant_convert
    except Exception:
        pass
except ImportError:
    print("\nERROR: openpyxl is not installed.")
    print("Fix: open Command Prompt and run:")
    print("  pip install openpyxl")
    input("\nPress Enter to close...")
    sys.exit(1)

# ── PANEL MAP (key → panel_id, export_label, folder_name) ────────────────────
PANEL_MAP = {
    "EX201_ML":           ("2-0",  "Import Declaration Report",                             "EX201_ML"),
    "EX202A_Release":     ("3-0",  "Release Goods from DZ Report",                          "EX202A_Release_Goods_From_DZ"),
    "EX202A_Consumption": ("3-1",  "Consumption of Goods Within DZ Report",                 "EX202A_Consumption_Within_DZ"),
    "EX202A_Enter":       ("3-2",  "Enter Goods into DZ Report",                            "EX202A_Enter_Goods_Into_DZ"),
    "EX202A_Transfer":    ("3-3",  "Transfer Goods to Another DZ Report",                   "EX202A_Transfer_To_Another_DZ"),
    "EX202A_Export":      ("3-4",  "Transfer Goods for Export From DZ Report",               "EX202A_Transfer_For_Export_From_DZ"),
    "EX202A_Import":      ("3-5",  "Import to DZ No Customs Check Report",                  "EX202A_Import_To_DZ"),
    "EX202A_Production":  ("3-6",  "Production Within DZ Report",                           "EX202A_Production_Within_DZ"),
    "EX202B":             ("2-1",  "Producer Declaration Report",                           "EX202B"),
    "EX203_ML":           ("2-2",  "Deductible Excise Tax Form Report",                     "EX203_ML"),
    "EX203A":             ("2-3",  "Local Purchase Form EX203A Report",                     "EX203A"),
    "EX203B":             ("3-7",  "Lost and Damaged Declaration Report",                   "EX203B"),
    "EX203C":             ("3-8",  "Transfer of Ownership within DZ Report",                "EX203C"),
    "EX203D":             ("2-4",  "Stockpile Declaration Report",                          "EX203D"),
    "EX203F":             ("3-9",  "TOO DZ Reg Seller to Non-Reg Purchaser Report",         "EX203F"),
    "EX203G":             ("3-11", "TOO DZ Non-Reg Seller to Reg Purchaser Report",         "EX203G"),
    "EX203H":             ("2-5",  "Local Purchase Form EX203H Report",                     "EX203H"),
    "EX204":              ("3-10", "Opening Stock Declaration Report",                      "EX204"),
    "AdminExceptions":    ("4-0",  "Administrative Exceptions Report",                      "AdminExceptions"),
}

DECL_LABELS = {
    "EX201_ML":           "EX201 - Import Declaration (Mainland)",
    "EX202A_Release":     "EX202A - Release Goods from DZ",
    "EX202A_Consumption": "EX202A - Consumption of Goods Within DZ",
    "EX202A_Enter":       "EX202A - Enter Goods into DZ",
    "EX202A_Transfer":    "EX202A - Transfer Goods to Another DZ",
    "EX202A_Export":      "EX202A - Transfer Goods for Export From DZ",
    "EX202A_Import":      "EX202A - Import to DZ",
    "EX202A_Production":  "EX202A - Production Within DZ",
    "EX202B":             "EX202B - Producer Declaration",
    "EX203_ML":           "EX203 - Deductible Excise Tax (Mainland)",
    "EX203A":             "EX203A - Local Purchase Form",
    "EX203B":             "EX203B - Lost and Damaged",
    "EX203C":             "EX203C - Transfer of Ownership DZ",
    "EX203D":             "EX203D - Stockpile Declaration",
    "EX203F":             "EX203F - TOO DZ Reg to Non-Reg",
    "EX203G":             "EX203G - TOO DZ Non-Reg to Reg",
    "EX203H":             "EX203H - Local Purchase (Mainland)",
    "EX204":              "EX204 - Opening Stock",
    "AdminExceptions":    "Administrative Exceptions",
}

YEAR_OPTIONS = ["2017", "2018", "2019", "2020", "2021", "2022", "2023", "2024", "2025", "2026", "2027"]

# ── COLORS / THEME (Andersen Consulting dark red + white) ─────────────────────
BG           = "#8B1A2B"
BG_CARD      = "#7A1525"
BG_INPUT     = "#6B1220"
FG           = "#FFFFFF"
FG_DIM       = "#E8C4C9"
ACCENT       = "#FFFFFF"
ACCENT_HOVER = "#F0D0D5"
SUCCESS      = "#7FD17F"
WARNING      = "#FFD580"
ERROR        = "#FF9999"
BORDER       = "#A03040"

# ══════════════════════════════════════════════════════════════════════════════
# JAVASCRIPT — Panel Navigation
# ══════════════════════════════════════════════════════════════════════════════

def js_click_panel(panel_id):
    return f"""
    () => {{
        var bdi = document.querySelector('[id*="panelScrollcontainer-{panel_id}-BDI-content"]');
        if (!bdi) return 'not found';
        var btnId = bdi.closest('[data-sap-ui]').id;
        var sapBtn = sap.ui.getCore().byId(btnId);
        if (sapBtn && sapBtn.firePress) {{
            sapBtn.firePress();
            return 'pressed: ' + btnId;
        }}
        return 'no sap button';
    }}
    """

JS_WAIT_FOR_TABLE = """
() => {
    // Still loading
    var b = document.querySelector('.sapUiLocalBusyIndicatorAnimation');
    if (b && b.getBoundingClientRect().width > 0) return 'not found';

    // Table must be visible
    var tableFound = false;
    var tables = document.querySelectorAll("table[id*='_Table-listUl'], table[id*='_List_table-listUl'], table[id*='-listUl']");
    for (var t = 0; t < tables.length; t++) {
        var r = tables[t].getBoundingClientRect();
        if (r.width > 0 && r.height > 0 && tables[t].id) { tableFound = true; break; }
    }
    if (!tableFound) {
        var allTables = document.querySelectorAll('table');
        for (var t = 0; t < allTables.length; t++) {
            var r = allTables[t].getBoundingClientRect();
            if (r.width > 0 && r.height > 0 && allTables[t].querySelector('th.sapMListTblHeaderCell')) { tableFound = true; break; }
        }
    }
    if (!tableFound) return 'not found';

    // Filter bar must also be ready — search input OR status combo visible
    var inputs = document.querySelectorAll('input[type="search"]');
    for (var i = 0; i < inputs.length; i++) {
        if (inputs[i].getBoundingClientRect().width > 0) return 'found';
    }
    var arrows = document.querySelectorAll('span[id$="_combobox-arrow"]');
    for (var i = 0; i < arrows.length; i++) {
        if (arrows[i].getBoundingClientRect().width > 0) return 'found';
    }
    return 'not found';
}
"""

JS_NAVIGATE_BACK = """
() => {
    window.focus();
    var link = Array.from(document.querySelectorAll('a.sapMLnk')).find(function(el) {
        return el.textContent.trim() === 'Excise Tax';
    });
    if (link) {
        link.click();
        return 'CLICKED';
    }
    return 'NOT_FOUND';
}
"""

# ══════════════════════════════════════════════════════════════════════════════
# JAVASCRIPT — ApplyFilters
# ══════════════════════════════════════════════════════════════════════════════

def js_search(search_term):
    return f"""
    () => {{
        // Derive active view prefix from the known table ID so we never
        // accidentally target a hidden panel's search field.
        var tableId = String(window.__PAD_TABLE_ID || '');
        var viewPrefix = '';
        if (tableId) {{
            var dash = tableId.indexOf('--');
            if (dash > -1) viewPrefix = tableId.substring(0, dash + 2); // e.g. "__xmlview19--"
        }}
        var all = document.querySelectorAll('input[type="search"]');
        var el = null;
        // Pass 1: same-view _searchField-I
        for (var i = 0; i < all.length; i++) {{
            var id = all[i].id;
            if (viewPrefix && id.indexOf(viewPrefix) === -1) continue;
            if (id.indexOf('_searchField-I') > -1 && all[i].getBoundingClientRect().width > 0) {{
                el = all[i]; break;
            }}
        }}
        // Pass 2: same-view Search-I (excludes searchbar)
        if (!el) {{
            for (var i = 0; i < all.length; i++) {{
                var id = all[i].id;
                if (viewPrefix && id.indexOf(viewPrefix) === -1) continue;
                if (id.indexOf('Search-I') > -1 && id.indexOf('searchbar') === -1 && all[i].getBoundingClientRect().width > 0) {{
                    el = all[i]; break;
                }}
            }}
        }}
        // Pass 3: any visible input in same view (covers searchbar-style IDs)
        if (!el) {{
            for (var i = 0; i < all.length; i++) {{
                var id = all[i].id;
                if (viewPrefix && id.indexOf(viewPrefix) === -1) continue;
                if (all[i].getBoundingClientRect().width > 0) {{ el = all[i]; break; }}
            }}
        }}
        // Pass 4: last resort — any visible search input (no view restriction)
        if (!el) {{
            for (var i = 0; i < all.length; i++) {{
                if (all[i].getBoundingClientRect().width > 0) {{ el = all[i]; break; }}
            }}
        }}
        if (!el) return 'FAIL';
        var sapId = el.id.replace(/-I$/, '');
        var ctrl = sap.ui.getCore().byId(sapId);
        if (!ctrl) return 'FAIL';
        ctrl.setValue('{search_term}');
        return ctrl.getValue();
    }}
    """

def js_verify_search(search_term):
    return f"""
    () => {{
        var tableId = String(window.__PAD_TABLE_ID || '');
        var viewPrefix = '';
        if (tableId) {{
            var dash = tableId.indexOf('--');
            if (dash > -1) viewPrefix = tableId.substring(0, dash + 2);
        }}
        var all = document.querySelectorAll('input[type="search"]');
        for (var i = 0; i < all.length; i++) {{
            var id = all[i].id;
            if (viewPrefix && id.indexOf(viewPrefix) === -1) continue;
            if (all[i].getBoundingClientRect().width > 0) return all[i].value;
        }}
        for (var i = 0; i < all.length; i++) {{
            if (all[i].getBoundingClientRect().width > 0) return all[i].value;
        }}
        return 'EMPTY';
    }}
    """

JS_SET_STATUS_APPROVED = """
() => {
    var tableId = String(window.__PAD_TABLE_ID || '');
    var viewPrefix = '';
    if (tableId) { var d = tableId.indexOf('--'); if (d > -1) viewPrefix = tableId.substring(0, d + 2); }

    function trySetApproved(arrowEl) {
        var comboId = arrowEl.id.replace('-arrow', '');
        var combo = sap.ui.getCore().byId(comboId);
        if (!combo) return null;
        var items = combo.getItems();
        for (var j = 0; j < items.length; j++) {
            if (items[j].getText().trim() === 'Approved') {
                combo.setSelectedKey(items[j].getKey());
                combo.setSelectedItem(items[j]);
                combo.setValue('Approved');
                return 'APPROVED_SET';
            }
        }
        if (items.length > 0) return 'NO_APPROVED';
        return null;
    }

    var arrows = document.querySelectorAll('span[id$="_combobox-arrow"]');

    // Pass 1: known status keyword patterns
    for (var i = 0; i < arrows.length; i++) {
        var id = arrows[i].id;
        if (viewPrefix && id.indexOf(viewPrefix) === -1) continue;
        if ((id.indexOf('Status_combobox') > -1 || id.indexOf('DecStatus_combobox') > -1 ||
             id.indexOf('myDecStatus_combobox') > -1 || id.indexOf('myDeclStatus_combobox') > -1) &&
            arrows[i].getBoundingClientRect().width > 0) {
            var r = trySetApproved(arrows[i]);
            if (r) return r;
        }
    }

    // Pass 2: any visible combo in same view that has an Approved item
    for (var i = 0; i < arrows.length; i++) {
        var id = arrows[i].id;
        if (viewPrefix && id.indexOf(viewPrefix) === -1) continue;
        if (id.indexOf('perpage') > -1) continue;  // skip page-size combo
        if (arrows[i].getBoundingClientRect().width > 0) {
            var r = trySetApproved(arrows[i]);
            if (r) return r;
        }
    }

    return 'ARROW_NOT_FOUND';
}
"""

JS_SET_STATUS_WAREHOUSE = """
() => {
    var tableId = String(window.__PAD_TABLE_ID || '');
    var viewPrefix = '';
    if (tableId) { var d = tableId.indexOf('--'); if (d > -1) viewPrefix = tableId.substring(0, d + 2); }

    function trySetWarehouse(arrowEl) {
        var comboId = arrowEl.id.replace('-arrow', '');
        var combo = sap.ui.getCore().byId(comboId);
        if (!combo) return null;
        var items = combo.getItems();
        for (var j = 0; j < items.length; j++) {
            var txt = items[j].getText().trim().toLowerCase();
            if (txt === 'approved by destination warehouse keeper' || txt === 'approved by warehouse keeper') {
                combo.setSelectedKey(items[j].getKey());
                combo.setSelectedItem(items[j]);
                combo.setValue(items[j].getText().trim());
                return 'WAREHOUSE_SET';
            }
        }
        if (items.length > 0) return 'NOT_FOUND';
        return null;
    }

    var arrows = document.querySelectorAll('span[id$="_combobox-arrow"]');

    // Pass 1: known status keyword patterns
    for (var i = 0; i < arrows.length; i++) {
        var id = arrows[i].id;
        if (viewPrefix && id.indexOf(viewPrefix) === -1) continue;
        if ((id.indexOf('Status_combobox') > -1 || id.indexOf('DecStatus_combobox') > -1 ||
             id.indexOf('myDecStatus_combobox') > -1 || id.indexOf('myDeclStatus_combobox') > -1) &&
            arrows[i].getBoundingClientRect().width > 0) {
            var r = trySetWarehouse(arrows[i]);
            if (r) return r;
        }
    }

    // Pass 2: any visible combo in same view that has a warehouse item
    for (var i = 0; i < arrows.length; i++) {
        var id = arrows[i].id;
        if (viewPrefix && id.indexOf(viewPrefix) === -1) continue;
        if (id.indexOf('perpage') > -1) continue;
        if (arrows[i].getBoundingClientRect().width > 0) {
            var r = trySetWarehouse(arrows[i]);
            if (r) return r;
        }
    }

    return 'FAIL';
}
"""

JS_SET_PAGE_1000 = """
() => {
    var tableId = String(window.__PAD_TABLE_ID || '');
    var viewPrefix = '';
    if (tableId) { var d = tableId.indexOf('--'); if (d > -1) viewPrefix = tableId.substring(0, d + 2); }
    var arrows = document.querySelectorAll('span[id*="perpage-arrow"][role="button"]');
    var arrow = null;
    for (var i = 0; i < arrows.length; i++) {
        if (viewPrefix && arrows[i].id.indexOf(viewPrefix) === -1) continue;
        if (arrows[i].getBoundingClientRect().width > 0) {
            arrow = arrows[i];
            break;
        }
    }
    if (!arrow) return 'FAIL';
    var comboId = arrow.id.replace('-arrow', '');
    var combo = sap.ui.getCore().byId(comboId);
    if (!combo) return 'FAIL';
    var items = combo.getItems();
    for (var j = 0; j < items.length; j++) {
        if (items[j].getText().trim() === '1000') {
            combo.setSelectedKey(items[j].getKey());
            combo.setSelectedItem(items[j]);
            combo.setValue('1000');
            return combo.getValue();
        }
    }
    return 'FAIL';
}
"""

JS_CLICK_GO = """
() => {
    var tableId = String(window.__PAD_TABLE_ID || '');
    var viewPrefix = '';
    if (tableId) { var d = tableId.indexOf('--'); if (d > -1) viewPrefix = tableId.substring(0, d + 2); }
    // Pass 1: visible Go button in same view
    var buttons = document.querySelectorAll('button');
    for (var i = 0; i < buttons.length; i++) {
        if (viewPrefix && buttons[i].id.indexOf(viewPrefix) === -1) continue;
        var bdi = buttons[i].querySelector('bdi');
        if (bdi && bdi.textContent.trim() === 'Go' && buttons[i].getBoundingClientRect().width > 0) {
            var sapBtn = sap.ui.getCore().byId(buttons[i].id);
            if (sapBtn && sapBtn.firePress) { sapBtn.firePress(); return 'SUCCESS'; }
            buttons[i].click(); return 'CLICKED';
        }
    }
    // Pass 2: Go button in toolbar overflow (hidden) — fire via SAP API using known ID pattern
    if (tableId) {
        var sapTableId = tableId.replace('-listUl', '');
        var goId = sapTableId + '_Go';
        var goBtn = sap.ui.getCore().byId(goId);
        if (goBtn && goBtn.firePress) { goBtn.firePress(); return 'SUCCESS_OVERFLOW'; }
    }
    // Pass 3: any Go button in same view, visible or not
    for (var i = 0; i < buttons.length; i++) {
        if (viewPrefix && buttons[i].id.indexOf(viewPrefix) === -1) continue;
        var bdi = buttons[i].querySelector('bdi');
        if (bdi && bdi.textContent.trim() === 'Go') {
            var sapBtn = sap.ui.getCore().byId(buttons[i].id);
            if (sapBtn && sapBtn.firePress) { sapBtn.firePress(); return 'SUCCESS_HIDDEN'; }
        }
    }
    return 'FAIL';
}
"""

JS_CHECK_NO_DATA = """
() => {
    // 1. Still loading — keep waiting
    var busy = document.querySelector('.sapUiLocalBusyIndicatorAnimation');
    if (busy && busy.getBoundingClientRect().width > 0) return 'NO_DATA';

    // 2. Explicit "No records found" cell
    var noData = document.querySelectorAll("td[id*='nodata-text']");
    for (var i = 0; i < noData.length; i++) {
        var rect = noData[i].getBoundingClientRect();
        if (rect.width > 0 && rect.height > 0) {
            if (noData[i].textContent.indexOf("No records found") > -1) return "NO_RECORDS";
            return "NO_DATA";
        }
    }

    // 3. Confirm rows are actually in the SAP binding — not just "busy gone"
    var tableId = String(window.__PAD_TABLE_ID || '');
    if (tableId) {
        var sapTableId = tableId.replace('-listUl', '');
        var sapTable = sap.ui.getCore().byId(sapTableId);
        if (sapTable) {
            var binding = sapTable.getBinding('items');
            if (binding && typeof binding.getLength === 'function') {
                var len = binding.getLength();
                if (len > 0) return 'HAS_DATA';
                return 'NO_DATA';
            }
        }
    }

    // 4. Fallback — check for visible data rows in DOM
    var rows = document.querySelectorAll('tr.sapMLIBActive, tr.sapMListTblRow');
    for (var r = 0; r < rows.length; r++) {
        var rr = rows[r].getBoundingClientRect();
        if (rr.width > 0 && rr.height > 0) return 'HAS_DATA';
    }
    return 'NO_DATA';
}
"""

# ══════════════════════════════════════════════════════════════════════════════
# JAVASCRIPT — Downloader
# ══════════════════════════════════════════════════════════════════════════════

JS_FIND_TABLE = """
() => {
    var b = document.querySelector('.sapUiLocalBusyIndicatorAnimation');
    if (b && b.getBoundingClientRect().width > 0) return 'TABLE_NOT_FOUND';
    function _visibleTable(t) {
        var r = t.getBoundingClientRect();
        return r.width > 0 && r.height > 0;
    }
    // Pass 1: known SAP list-table ID patterns
    var tables = document.querySelectorAll("table[id*='_Table-listUl'], table[id*='_List_table-listUl'], table[id*='-listUl']");
    for (var t = 0; t < tables.length; t++) {
        if (_visibleTable(tables[t]) && tables[t].id) {
            window.__PAD_TABLE_ID = tables[t].id;
            return tables[t].id;
        }
    }
    // Pass 2: any visible SAP table (has sapMListTblHeaderCell th) — catches ETP/myDec_LIST style IDs
    var allTables = document.querySelectorAll("table");
    for (var t = 0; t < allTables.length; t++) {
        if (!_visibleTable(allTables[t]) || !allTables[t].id) continue;
        if (allTables[t].querySelector("th.sapMListTblHeaderCell")) {
            window.__PAD_TABLE_ID = allTables[t].id;
            return allTables[t].id;
        }
    }
    // Pass 3: header text fallback
    for (var t = 0; t < allTables.length; t++) {
        if (!_visibleTable(allTables[t])) continue;
        var headers = allTables[t].querySelectorAll("th");
        for (var h = 0; h < headers.length; h++) {
            var txt = headers[h].innerText.trim();
            if (txt === "Transaction Number" || txt === "Excise Tax Period" || txt === "Declaration Number") {
                window.__PAD_TABLE_ID = allTables[t].id;
                return allTables[t].id;
            }
        }
    }
    return "TABLE_NOT_FOUND";
}
"""

JS_GET_ROW_COUNT = """
() => {
    var tableId = String(window.__PAD_TABLE_ID || "");
    if (!tableId) return "0";
    var sapTableId = tableId.replace("-listUl", "");
    var sapTable = sap.ui.getCore().byId(sapTableId);
    // Primary: SAP binding length — works for all declaration types
    if (sapTable) {
        var binding = sapTable.getBinding('items');
        if (binding && typeof binding.getLength === 'function') {
            var len = binding.getLength();
            if (len > 0) return String(len);
        }
    }
    // Fallback: row-count toolbar span
    var rowCountSpan = document.getElementById(sapTableId + "_rowCount");
    if (rowCountSpan) {
        var text = String(rowCountSpan.innerText || rowCountSpan.textContent || "");
        var m = text.match(/of\\s+([\\d,]+)\\s+records?/i);
        if (!m) m = text.match(/([\\d,]+)/);
        if (m) return String(m[1]).replace(/,/g, "");
    }
    // Last resort: count rendered SAP items
    if (sapTable && sapTable.getItems) {
        var n = sapTable.getItems().length;
        if (n > 0) return String(n);
    }
    return "0";
}
"""

JS_GET_PAGE_SIZE = """
() => {
    var table = document.getElementById(window.__PAD_TABLE_ID);
    if (!table) return "100";
    var sapTableId = table.id.replace("-listUl", "");
    var sapTable = sap.ui.getCore().byId(sapTableId);
    if (sapTable && sapTable.getItems) {
        return String(sapTable.getItems().length);
    }
    var rows = Array.from(table.querySelectorAll("tr")).filter(r => r.querySelector("td"));
    return String(rows.length);
}
"""

JS_CLEAR_POPUPS = """
() => {
    // Close open SAP popups via their own API only — document.body.click() would
    // trigger SAP's filter-bar outside-click handler and reset the status combo.
    var openPopups = document.querySelectorAll("div[class*='sapMPopover'], div[class*='sapMDialog'], div[class*='sapUiPopup'], div[class*='sapMActionSheet']");
    for (var p = 0; p < openPopups.length; p++) {
        var rect = openPopups[p].getBoundingClientRect();
        if (rect.width > 0 && rect.height > 0) {
            var sapId = openPopups[p].id;
            if (sapId) {
                var ctrl = sap.ui.getCore().byId(sapId);
                if (ctrl && ctrl.close) ctrl.close();
            }
        }
    }
    return "CLEARED";
}
"""

JS_CLICK_EXPORT = """
() => {
    const LABEL = "Export to Excel";
    var allBdi = Array.from(document.querySelectorAll("bdi"));
    var textNode = null;
    for (var i = 0; i < allBdi.length; i++) {
        if (allBdi[i].textContent.trim() === LABEL) {
            var rect = allBdi[i].getBoundingClientRect();
            if (rect.width > 0 && rect.height > 0) {
                textNode = allBdi[i];
                break;
            }
        }
    }
    if (!textNode) return "TEXT_NOT_FOUND";
    var domButton = textNode.closest("button");
    if (!domButton || !domButton.id) return "BUTTON_NOT_FOUND";
    var btn = sap.ui.getCore().byId(domButton.id);
    if (btn && btn.firePress) {
        btn.firePress();
        return "EXPORT_TRIGGERED";
    }
    domButton.click();
    return "EXPORT_CLICKED";
}
"""

JS_CLICK_NEXT = """
() => {
    var nextBtn = null;
    var candidates = document.querySelectorAll("button[id*='_Table_next'], span[id*='_Table_next-inner']");
    if (candidates.length > 0) nextBtn = candidates[0];
    if (!nextBtn) return "NEXT_NOT_FOUND";
    var btn = nextBtn.closest("button");
    if (btn) {
        if (btn.disabled || btn.classList.contains("sapMBtnDisabled")) return "NEXT_DISABLED";
        btn.click();
        return "NEXT_CLICKED";
    }
    nextBtn.click();
    return "NEXT_CLICKED_SPAN";
}
"""


def js_scroll_to_row(idx):
    return f"""
    () => {{
        var rowIndex = {idx};
        var table = document.getElementById(window.__PAD_TABLE_ID);
        if (!table) return "TABLE_NOT_FOUND";
        var sapTableId = table.id.replace("-listUl", "");
        var sapTable = sap.ui.getCore().byId(sapTableId);
        if (sapTable && sapTable.getItems) {{
            var items = sapTable.getItems();
            if (rowIndex < items.length) {{
                var domRef = items[rowIndex].getDomRef();
                if (domRef) {{
                    domRef.scrollIntoView({{behavior: "instant", block: "center"}});
                    return "SCROLLED_TO_ROW_" + rowIndex;
                }}
                if (sapTable.scrollToIndex) {{
                    sapTable.scrollToIndex(rowIndex);
                    return "SCROLLED_VIA_API_" + rowIndex;
                }}
            }}
            return "ROW_NOT_IN_ITEMS";
        }}
        var rows = Array.from(table.querySelectorAll("tr")).filter(r => r.querySelector("td"));
        if (rowIndex < rows.length) {{
            rows[rowIndex].scrollIntoView({{behavior: "instant", block: "center"}});
            return "SCROLLED_DOM_" + rowIndex;
        }}
        return "ROW_NOT_FOUND";
    }}
    """


def js_extract_txn(idx):
    return f"""
    () => {{
        var rowIndex = {idx};
        var table = document.getElementById(window.__PAD_TABLE_ID);
        if (!table) return "TABLE_NOT_FOUND";
        var tRect = table.getBoundingClientRect();
        if (tRect.width === 0 || tRect.height === 0) return "TABLE_NOT_VISIBLE";
        var headers = table.querySelectorAll("th");
        var columnIndex = -1;
        for (var i = 0; i < headers.length; i++) {{
            if (headers[i].innerText.trim() === "Transaction Number") {{
                columnIndex = i;
                break;
            }}
        }}
        if (columnIndex === -1) return "COLUMN_NOT_FOUND";
        var sapTableId = table.id.replace("-listUl", "");
        var sapTable = sap.ui.getCore().byId(sapTableId);
        if (sapTable && sapTable.getItems) {{
            var items = sapTable.getItems();
            if (rowIndex < items.length) {{
                var ctx = items[rowIndex].getBindingContext();
                if (ctx) {{
                    var obj = ctx.getObject();
                    for (var key in obj) {{
                        if (key.toLowerCase().indexOf("transaction") >= 0 || key.toLowerCase().indexOf("txn") >= 0) {{
                            return String(obj[key]);
                        }}
                    }}
                }}
            }}
        }}
        var rows = Array.from(table.querySelectorAll("tr")).filter(r => r.querySelector("td"));
        if (rowIndex >= rows.length) return "END";
        var cells = rows[rowIndex].querySelectorAll("td");
        if (columnIndex >= cells.length) return "EMPTY";
        return cells[columnIndex].innerText.trim();
    }}
    """


def js_click_more(idx):
    return f"""
    () => {{
        var rowIndex = {idx};
        var table = document.getElementById(window.__PAD_TABLE_ID);
        if (!table) return "TABLE_NOT_FOUND";
        var sapTableId = table.id.replace("-listUl", "");
        var sapTable = sap.ui.getCore().byId(sapTableId);
        if (sapTable && sapTable.getItems) {{
            var items = sapTable.getItems();
            if (rowIndex < items.length) {{
                var domRef = items[rowIndex].getDomRef();
                if (domRef) {{
                    var moreBtn = domRef.querySelector("span[title='More'], [title='More'], button[aria-label='More'], span[aria-label='More']");
                    if (moreBtn) {{
                        moreBtn.click();
                        return "clicked item " + rowIndex;
                    }}
                }}
            }}
        }}
        var rows = Array.from(table.querySelectorAll("tr")).filter(r => r.querySelector("td"));
        if (rowIndex >= rows.length) return "END";
        var row = rows[rowIndex];
        var moreBtn = row.querySelector("span[title='More'], [title='More'], button[aria-label='More'], span[aria-label='More']");
        if (!moreBtn) return "MORE_NOT_FOUND";
        moreBtn.click();
        return "clicked row " + rowIndex;
    }}
    """


# ── HELPERS ───────────────────────────────────────────────────────────────────

def _list_downloads(directory):
    """All completed (non-temp) files in the downloads folder."""
    all_files = glob.glob(os.path.join(directory, "*"))
    return [
        f for f in all_files
        if os.path.isfile(f)
        and not f.endswith(".crdownload")
        and not f.endswith(".tmp")
        and not f.endswith(".partial")
        and not os.path.basename(f).startswith("~$")
        and not os.path.basename(f).startswith(".")
    ]

def _wait_for_stable_file(filepath, stable_secs=1.0):
    """Wait until a file's size stops changing — means download is complete."""
    prev_size = -1
    stable_count = 0
    for _ in range(40):  # max 20s
        try:
            size = os.path.getsize(filepath)
        except OSError:
            time.sleep(0.5)
            continue
        if size > 0 and size == prev_size:
            stable_count += 1
            if stable_count >= 2:  # stable for 2 checks = ~1s
                return True
        else:
            stable_count = 0
        prev_size = size
        time.sleep(0.5)
    return False

def count_xlsx(directory):
    return len(_list_downloads(directory))

def get_latest_xlsx(directory):
    files = _list_downloads(directory)
    if not files:
        return None
    return max(files, key=os.path.getmtime)


# ══════════════════════════════════════════════════════════════════════════════
# GUI APP
# ══════════════════════════════════════════════════════════════════════════════

class ExciseScraperApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("FTA Excise Portal Scraper")
        self.root.configure(bg=BG)
        self.root.resizable(True, True)
        self.root.minsize(600, 700)
        self._center(740, 860)

        self.is_running = False
        self.stop_requested = False
        self.pw_page = None
        self.pw_browser = None
        self.pw_instance = None
        self._chrome_proc = None

        # Single persistent thread for ALL Playwright calls
        self._pw_queue = queue.Queue()
        self._pw_thread = threading.Thread(target=self._pw_worker, daemon=True)
        self._pw_thread.start()

        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.mainloop()

    def _pw_worker(self):
        """Single thread that executes all Playwright tasks from the queue."""
        while True:
            fn = self._pw_queue.get()
            if fn is None:
                break
            try:
                fn()
            except Exception as e:
                err = str(e)
                self.root.after(0, lambda err=err: self._log(f"Worker error: {err}", "error"))
                # Reset running state so the UI doesn't stay locked after a crash
                self.root.after(0, self._reset_after_crash)
            finally:
                self._pw_queue.task_done()

    def _reset_after_crash(self):
        """Restore UI to idle state after an unhandled worker exception."""
        self.is_running = False
        self.stop_requested = False
        self._stop_snail()
        self.start_btn.configure(state="normal" if self.pw_page else "disabled")
        self.stop_btn.configure(state="disabled", text="Stop")
        self.open_btn.configure(state="normal")
        self.status_var.set("Stopped — an error occurred, check the log")

    def _center(self, w, h):
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    def _build_ui(self):
        # ── Header ──
        header = tk.Frame(self.root, bg=BG, height=80)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="FTA Excise Portal Scraper",
                 font=("Helvetica Neue", 22, "bold"), fg=ACCENT, bg=BG).pack(pady=(20, 2))
        tk.Label(header, text="Automated declaration export tool",
                 font=("Helvetica Neue", 11), fg=FG_DIM, bg=BG).pack()

        tk.Frame(self.root, bg=BORDER, height=1).pack(fill="x", padx=30, pady=(10, 0))

        # ── Settings Card ──
        card = tk.Frame(self.root, bg=BG_CARD, highlightbackground=BORDER, highlightthickness=1)
        card.pack(fill="x", padx=30, pady=10)

        tk.Label(card, text="SETTINGS", font=("Helvetica Neue", 10, "bold"),
                 fg=FG_DIM, bg=BG_CARD).pack(anchor="w", padx=20, pady=(12, 6))

        # ── Date Mode ──
        dm_frame = tk.Frame(card, bg=BG_CARD)
        dm_frame.pack(fill="x", padx=20, pady=4)
        tk.Label(dm_frame, text="Date Range", font=("Helvetica Neue", 11),
                 fg=FG, bg=BG_CARD, width=14, anchor="w").pack(side="left")

        # ── Date Range Inputs ──
        MONTHS = ["January","February","March","April","May","June",
                  "July","August","September","October","November","December"]
        YEARS = ["2017","2018","2019","2020","2021","2022","2023","2024","2025","2026","2027"]
        self.range_frame = tk.Frame(card, bg=BG_CARD)
        self.range_frame.pack(fill="x", padx=20, pady=4)
        tk.Label(self.range_frame, text="From", font=("Helvetica Neue", 11),
                 fg=FG, bg=BG_CARD, width=14, anchor="w").pack(side="left")
        self.range_start_month = tk.StringVar(value="January")
        self.range_start_year  = tk.StringVar(value="2025")
        self.range_end_month   = tk.StringVar(value="December")
        self.range_end_year    = tk.StringVar(value="2025")
        for var, opts in [(self.range_start_month, MONTHS),(self.range_start_year, YEARS)]:
            om = tk.OptionMenu(self.range_frame, var, *opts)
            om.configure(bg=BG_INPUT, fg=FG, activebackground=BG_INPUT,
                         activeforeground=FG, highlightthickness=0, relief="flat",
                         font=("Helvetica Neue", 10))
            om.pack(side="left", padx=2)
        tk.Label(self.range_frame, text="To", font=("Helvetica Neue", 11),
                 fg=FG, bg=BG_CARD).pack(side="left", padx=(10,0))
        for var, opts in [(self.range_end_month, MONTHS),(self.range_end_year, YEARS)]:
            om = tk.OptionMenu(self.range_frame, var, *opts)
            om.configure(bg=BG_INPUT, fg=FG, activebackground=BG_INPUT,
                         activeforeground=FG, highlightthickness=0, relief="flat",
                         font=("Helvetica Neue", 10))
            om.pack(side="left", padx=2)
        tk.Button(self.range_frame, text="Last Month", command=self._set_last_month,
                  bg=BG_INPUT, fg=ACCENT, relief="flat", padx=8,
                  font=("Helvetica Neue", 10), cursor="hand2").pack(side="left", padx=(10, 0))

        # ── Date range validation error label ──
        self.date_error_label = tk.Label(card, text="", fg="#FF6B6B", bg=BG_CARD,
                                          font=("Helvetica Neue", 9, "italic"))
        # Attach validation traces
        for v in [self.range_start_month, self.range_start_year,
                  self.range_end_month, self.range_end_year]:
            v.trace_add("write", lambda *_: self.root.after(0, self._validate_date_range))

        # ── Declaration Types — checkbox list ──
        decl_header = tk.Frame(card, bg=BG_CARD)
        decl_header.pack(fill="x", padx=20, pady=(8, 2))
        tk.Label(decl_header, text="Declaration Types", font=("Helvetica Neue", 11),
                 fg=FG, bg=BG_CARD, anchor="w").pack(side="left")
        sel_frame = tk.Frame(decl_header, bg=BG_CARD)
        sel_frame.pack(side="right")
        tk.Button(sel_frame, text="Select All", command=self._select_all_decl,
                  bg=BG_INPUT, fg=ACCENT, relief="flat", padx=6,
                  font=("Helvetica Neue", 9), cursor="hand2").pack(side="left", padx=2)
        tk.Button(sel_frame, text="Clear All", command=self._clear_all_decl,
                  bg=BG_INPUT, fg=FG_DIM, relief="flat", padx=6,
                  font=("Helvetica Neue", 9), cursor="hand2").pack(side="left", padx=2)

        decl_container = tk.Frame(card, bg=BG_INPUT, highlightbackground=BORDER, highlightthickness=1)
        decl_container.pack(fill="x", padx=20, pady=(0, 6))
        decl_canvas = tk.Canvas(decl_container, bg=BG_INPUT, highlightthickness=0, height=130)
        decl_scrollbar = tk.Scrollbar(decl_container, orient="vertical", command=decl_canvas.yview)
        decl_inner = tk.Frame(decl_canvas, bg=BG_INPUT)
        decl_inner.bind("<Configure>", lambda e: decl_canvas.configure(scrollregion=decl_canvas.bbox("all")))
        decl_canvas.create_window((0, 0), window=decl_inner, anchor="nw")
        decl_canvas.configure(yscrollcommand=decl_scrollbar.set)
        decl_canvas.pack(side="left", fill="both", expand=True)
        decl_scrollbar.pack(side="right", fill="y")

        def _on_mousewheel(event):
            # macOS trackpad sends small deltas; Windows sends multiples of 120
            if abs(event.delta) < 10:
                decl_canvas.yview_scroll(-1 * event.delta, "units")
            else:
                decl_canvas.yview_scroll(-1 * (event.delta // 120), "units")
        decl_canvas.bind("<MouseWheel>", _on_mousewheel)
        decl_inner.bind("<MouseWheel>", _on_mousewheel)

        self.decl_vars = {}
        for key, label in DECL_LABELS.items():
            var = tk.BooleanVar(value=False)
            cb = tk.Checkbutton(decl_inner, text=label, variable=var,
                                bg=BG_INPUT, fg=FG, selectcolor=BG_CARD,
                                activebackground=BG_INPUT, activeforeground=ACCENT,
                                font=("Helvetica Neue", 10), anchor="w", highlightthickness=0, bd=0)
            cb.pack(fill="x", padx=8, pady=1)
            cb.bind("<MouseWheel>", _on_mousewheel)
            self.decl_vars[key] = var

        # ── Save Folder ──
        row_folder = tk.Frame(card, bg=BG_CARD)
        row_folder.pack(fill="x", padx=20, pady=(4, 12))
        tk.Label(row_folder, text="Save Folder", font=("Helvetica Neue", 11),
                 fg=FG, bg=BG_CARD, width=14, anchor="w").pack(side="left")
        default_dl = os.path.join(os.environ.get("USERPROFILE", os.path.expanduser("~")), "Downloads")
        self.folder_var = tk.StringVar(value=default_dl)
        tk.Entry(row_folder, textvariable=self.folder_var, width=30,
                 bg=BG_INPUT, fg=FG, insertbackground=FG, relief="flat",
                 font=("Helvetica Neue", 11)).pack(side="left", padx=(10, 5))
        tk.Button(row_folder, text="Browse", command=self._browse_folder,
                  bg=BG_INPUT, fg=FG, relief="flat", padx=8,
                  font=("Helvetica Neue", 10)).pack(side="left")

        # ── Buttons Row ──
        btn_frame = tk.Frame(self.root, bg=BG)
        btn_frame.pack(fill="x", padx=30, pady=(0, 8))

        self.open_btn = tk.Button(
            btn_frame, text="1. Open Browser & Login",
            font=("Helvetica Neue", 12, "bold"), bg="#FFFFFF", fg="#8B1A2B",
            activebackground="#F0D0D5", activeforeground="#8B1A2B",
            relief="flat", padx=20, pady=10, cursor="hand2",
            command=self._open_browser)
        self.open_btn.pack(side="left", expand=True, fill="x", padx=(0, 5))

        self.start_btn = tk.Button(
            btn_frame, text="2. Start Scraping",
            font=("Helvetica Neue", 12, "bold"), bg="#FFFFFF", fg="#2D6A2D",
            activebackground="#F0D0D5", activeforeground="#2D6A2D",
            relief="flat", padx=20, pady=10, cursor="hand2",
            command=self._start_scrape, state="disabled")
        self.start_btn.pack(side="left", expand=True, fill="x", padx=(5, 5))

        self.stop_btn = tk.Button(
            btn_frame, text="Stop",
            font=("Helvetica Neue", 12, "bold"), bg="#FFFFFF", fg="#C0392B",
            activebackground="#F0D0D5", activeforeground="#C0392B",
            relief="flat", padx=20, pady=10, cursor="hand2",
            command=self._stop_scrape, state="disabled")
        self.stop_btn.pack(side="left", padx=(5, 0))

        # ── Stats Card ──
        stats_card = tk.Frame(self.root, bg=BG_CARD, highlightbackground=BORDER, highlightthickness=1)
        stats_card.pack(fill="x", padx=30, pady=(0, 8))
        stats_inner = tk.Frame(stats_card, bg=BG_CARD)
        stats_inner.pack(fill="x", padx=20, pady=10)

        style = ttk.Style()
        style.theme_use('default')
        style.configure("Custom.Horizontal.TProgressbar",
                        troughcolor="#6B1220", background="#FFFFFF", thickness=8, borderwidth=0)
        self.progress_var = tk.DoubleVar(value=0)
        ttk.Progressbar(stats_inner, variable=self.progress_var, maximum=100,
                        style="Custom.Horizontal.TProgressbar").pack(fill="x", pady=(0, 8))

        stats_row = tk.Frame(stats_inner, bg=BG_CARD)
        stats_row.pack(fill="x")
        for col in range(4):
            stats_row.columnconfigure(col, weight=1)

        self.stat_total = self._stat_box(stats_row, "Total Rows", "0", 0)
        self.stat_downloaded = self._stat_box(stats_row, "Downloaded", "0", 1)
        self.stat_skipped = self._stat_box(stats_row, "Skipped", "0", 2)
        self.stat_progress = self._stat_box(stats_row, "Progress", "0%", 3)

        # ── Status ──
        self.status_var = tk.StringVar(value="Ready — configure settings and open browser")
        tk.Label(self.root, textvariable=self.status_var,
                 font=("Helvetica Neue", 11), fg=ACCENT, bg=BG).pack(pady=(0, 2))

        # ── Snail animation strip ──
        self.snail_canvas = tk.Canvas(self.root, bg=BG, height=28, highlightthickness=0)
        self.snail_canvas.pack(fill="x", padx=30)
        self.snail_item = self.snail_canvas.create_text(-30, 14, text="🐌",
                                                         font=("Helvetica Neue", 18), fill=ACCENT)
        self.snail_canvas.itemconfigure(self.snail_item, state="hidden")
        self._snail_x = -30
        self._snail_animating = False

        # ── Log Area ──
        tk.Label(self.root, text="ACTIVITY LOG", font=("Helvetica Neue", 10, "bold"),
                 fg=FG_DIM, bg=BG, anchor="w").pack(fill="x", padx=32, pady=(0, 3))
        log_frame = tk.Frame(self.root, bg=BG_CARD, highlightbackground=BORDER, highlightthickness=1)
        log_frame.pack(fill="both", expand=True, padx=30, pady=(0, 16))
        self.log_text = tk.Text(log_frame, bg="#6B1220", fg="#FFFFFF", font=("SF Mono", 10),
                                relief="flat", wrap="word", insertbackground=FG,
                                selectbackground=ACCENT, selectforeground=BG,
                                padx=12, pady=8, height=8)
        scrollbar = tk.Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True)
        self.log_text.configure(state="disabled")
        self.log_text.tag_configure("info", foreground=FG)
        self.log_text.tag_configure("success", foreground=SUCCESS)
        self.log_text.tag_configure("warning", foreground=WARNING)
        self.log_text.tag_configure("error", foreground=ERROR)
        self.log_text.tag_configure("accent", foreground="#FFB3BA")

        # ── Footer ──
        tk.Label(self.root, text="Report any errors to Garv — let's fix them soon",
                 font=("Helvetica Neue", 9), fg=FG_DIM, bg=BG).pack(pady=(4, 8))

    # ── UI Helpers ────────────────────────────────────────────────────────────

    def _stat_box(self, parent, label, value, col):
        frame = tk.Frame(parent, bg=BG_CARD)
        frame.grid(row=0, column=col, sticky="nsew", padx=5)
        tk.Label(frame, text=label, font=("Helvetica Neue", 9), fg=FG_DIM, bg=BG_CARD).pack()
        val_label = tk.Label(frame, text=value, font=("Helvetica Neue", 18, "bold"), fg=FG, bg=BG_CARD)
        val_label.pack()
        return val_label

    def _log(self, message, tag="info"):
        self.log_text.configure(state="normal")
        ts = time.strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{ts}] {message}\n", tag)
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _update_stats(self, total=None, downloaded=None, skipped=None, progress=None):
        if total is not None:
            self.stat_total.configure(text=str(total))
        if downloaded is not None:
            self.stat_downloaded.configure(text=str(downloaded))
        if skipped is not None:
            self.stat_skipped.configure(text=str(skipped))
        if progress is not None:
            self.stat_progress.configure(text=f"{progress}%")
            self.progress_var.set(progress)


    def _set_last_month(self):
        import datetime
        MONTHS = ["January","February","March","April","May","June",
                  "July","August","September","October","November","December"]
        today = datetime.date.today()
        first_of_this_month = today.replace(day=1)
        last_month = first_of_this_month - datetime.timedelta(days=1)
        m = MONTHS[last_month.month - 1]
        y = str(last_month.year)
        self.range_start_month.set(m)
        self.range_start_year.set(y)
        self.range_end_month.set(m)
        self.range_end_year.set(y)

    def _browse_folder(self):
        folder = filedialog.askdirectory(initialdir=self.folder_var.get())
        if folder:
            self.folder_var.set(folder)

    def _select_all_decl(self):
        for var in self.decl_vars.values():
            var.set(True)

    def _clear_all_decl(self):
        for var in self.decl_vars.values():
            var.set(False)

    def _validate_date_range(self):
        MONTHS = ["January","February","March","April","May","June",
                  "July","August","September","October","November","December"]
        try:
            si = MONTHS.index(self.range_start_month.get())
            sy = int(self.range_start_year.get())
            ei = MONTHS.index(self.range_end_month.get())
            ey = int(self.range_end_year.get())
        except (ValueError, Exception):
            return
        if (sy > ey) or (sy == ey and si > ei):
            self.date_error_label.configure(text="  From date can't be after To date")
            self.date_error_label.pack(fill="x", padx=20, pady=(0, 4))
            self.start_btn.configure(state="disabled")
        else:
            self.date_error_label.configure(text="")
            self.date_error_label.pack_forget()
            if self.pw_page and not self.is_running:
                self.start_btn.configure(state="normal")

    def _start_snail(self):
        self._snail_x = -30
        self._snail_animating = True
        self.snail_canvas.itemconfigure(self.snail_item, state="normal")
        self._animate_snail()

    def _stop_snail(self):
        self._snail_animating = False
        self.snail_canvas.itemconfigure(self.snail_item, state="hidden")

    def _animate_snail(self):
        if not self._snail_animating:
            return
        w = self.snail_canvas.winfo_width() or 680
        self._snail_x += 2
        if self._snail_x > w + 30:
            self._snail_x = -30
        self.snail_canvas.coords(self.snail_item, self._snail_x, 14)
        self.root.after(40, self._animate_snail)

    def _get_selected_decls(self):
        return [key for key, var in self.decl_vars.items() if var.get()]

    def _get_search_terms(self):
        """Return list of search terms based on date mode."""
        MONTHS = ["January","February","March","April","May","June",
                  "July","August","September","October","November","December"]
        sm = self.range_start_month.get()
        sy = int(self.range_start_year.get())
        em = self.range_end_month.get()
        ey = int(self.range_end_year.get())
        si = MONTHS.index(sm)
        ei = MONTHS.index(em)
        terms = []
        y, m = sy, si
        while (y < ey) or (y == ey and m <= ei):
            terms.append(f"{MONTHS[m]} {y}")
            m += 1
            if m > 11:
                m = 0
                y += 1
        return terms

    # ── Browser ───────────────────────────────────────────────────────────────

    def _open_browser(self):
        self.open_btn.configure(state="disabled", text="Connecting...")
        self._log("Checking for existing browser session...", "accent")
        self.status_var.set("Connecting to browser...")
        self._pw_queue.put(self._launch_browser)

    def _launch_browser(self):
        try:
            self.pw_instance = sync_playwright().start()
            dl = self.folder_var.get()
            os.makedirs(dl, exist_ok=True)

            # ── Step 1: Try to reuse an already-running Edge on port 9222 ──
            try:
                self.pw_browser = self.pw_instance.chromium.connect_over_cdp(
                    "http://localhost:9222", timeout=2000
                )
                contexts = self.pw_browser.contexts
                if contexts:
                    pages = contexts[0].pages
                    self.pw_page = pages[0] if pages else contexts[0].new_page()
                else:
                    context = self.pw_browser.new_context(accept_downloads=True)
                    self.pw_page = context.new_page()
                self.root.after(0, lambda: self._browser_ready(reconnected=True))
                return
            except Exception:
                pass  # no existing browser — launch a fresh one

            # ── Step 2: Launch Edge with a persistent profile so login is remembered ──
            edge_paths = [
                r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
                r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
            ]
            edge_bin = None
            for p in edge_paths:
                if os.path.exists(p):
                    edge_bin = p
                    break

            # Persistent profile dir — Edge saves cookies/session here across runs
            user_data = os.path.join(
                os.environ.get("LOCALAPPDATA", os.path.expanduser("~")),
                "ExciseScraper", "EdgeProfile"
            )

            if edge_bin:
                self._chrome_proc = subprocess.Popen([
                    edge_bin,
                    "--remote-debugging-port=9222",
                    f"--user-data-dir={user_data}",
                    "https://eservices.tax.gov.ae/#/Logon"
                ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                time.sleep(3)

                self.pw_browser = self.pw_instance.chromium.connect_over_cdp("http://localhost:9222")
                contexts = self.pw_browser.contexts
                if contexts:
                    pages = contexts[0].pages
                    self.pw_page = pages[0] if pages else contexts[0].new_page()
                else:
                    context = self.pw_browser.new_context(accept_downloads=True)
                    self.pw_page = context.new_page()
                    self.pw_page.goto("https://eservices.tax.gov.ae/#/Logon",
                                      timeout=300000, wait_until="domcontentloaded")
            else:
                # Fallback: Playwright bundled Edge
                self.pw_browser = self.pw_instance.chromium.launch(
                    channel="msedge", headless=False, downloads_path=dl
                )
                context = self.pw_browser.new_context(accept_downloads=True)
                self.pw_page = context.new_page()
                self.pw_page.goto("https://eservices.tax.gov.ae/#/Logon",
                                  timeout=300000, wait_until="domcontentloaded")

            self.root.after(0, lambda: self._browser_ready(reconnected=False))
        except Exception as e:
            self.root.after(0, lambda: self._browser_error(str(e)))

    def _browser_ready(self, reconnected=False):
        self.open_btn.configure(text="Browser Open", bg="#D4EDDA", fg="#2D6A2D")
        if reconnected:
            self._log("Reconnected to existing browser session — ready to scrape", "success")
            self.status_var.set("Reconnected to existing session — click 'Start Scraping'")
        else:
            self._log("Browser opened — log in, then click Start Scraping", "success")
            self.status_var.set("Log in to the portal, then click 'Start Scraping'")
        self._validate_date_range()  # only enable Start if date range is currently valid

    def _browser_error(self, msg):
        self.open_btn.configure(state="normal", text="1. Open Browser & Login")
        self._log(f"Browser error: {msg}", "error")
        self.status_var.set("Browser failed — check log")

    # ── Start / Stop ──────────────────────────────────────────────────────────

    def _start_scrape(self):
        if not self.pw_page:
            self._log("No browser open!", "error")
            return
        selected = self._get_selected_decls()
        if not selected:
            self._log("No declarations selected!", "error")
            return
        terms = self._get_search_terms()
        if not terms or (len(terms) == 1 and not terms[0]):
            self._log("No date/search term specified!", "error")
            return
        self.is_running = True
        self.stop_requested = False
        self.start_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self.open_btn.configure(state="disabled")
        self.status_var.set("Scraping in progress...")
        self._start_snail()
        self._pw_queue.put(self._scrape_main)

    def _stop_scrape(self):
        self.stop_requested = True
        self._stop_snail()
        self.stop_btn.configure(state="disabled", text="Stopping...")
        self._log("Stop requested — finishing current row...", "warning")

    def _scrape_done(self, downloaded, skipped, total):
        self.is_running = False
        self.stop_requested = False
        self._stop_snail()
        self.start_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled", text="Stop")
        self.status_var.set(f"Done! {downloaded} downloaded, {skipped} skipped out of {total}")

    # ══════════════════════════════════════════════════════════════════════════
    # MAIN SCRAPE LOOP — mirrors the PAD Main flow
    # ══════════════════════════════════════════════════════════════════════════

    def _scrape_main(self):
        page = self.pw_page
        base_dir = self.folder_var.get()

        # Run folder created on-demand when first file is downloaded
        run_stamp = time.strftime("%Y-%m-%d_%H-%M")
        download_dir = os.path.join(base_dir, run_stamp)
        self.root.after(0, lambda d=download_dir: self._log(f"Run folder: {d}", "accent"))

        selected_decls = self._get_selected_decls()
        search_terms = self._get_search_terms()

        grand_downloaded = 0
        grand_skipped = 0
        grand_total = 0

        for decl_idx, decl_key in enumerate(selected_decls):
            if self.stop_requested:
                break

            panel_id, export_label, folder_name = PANEL_MAP[decl_key]
            label = DECL_LABELS[decl_key]

            # Loop each month in the date range for this declaration
            panel_open = False  # track whether we've already clicked into this panel
            for term_idx, search_term in enumerate(search_terms):
                if self.stop_requested:
                    break

                dest_folder = os.path.join(download_dir, folder_name, search_term)
                # folder created on-demand in _download_rows, not here

                self.root.after(0, lambda l=label, st=search_term: self._log(
                    f"=== {l} | {st} ===", "accent"))
                self.root.after(0, lambda l=label, st=search_term: self.status_var.set(
                    f"Processing {l} — {st}"))

                if not panel_open:
                    # ── 1. Click panel — retry until it registers ──
                    self.root.after(0, lambda: self._log("Clicking panel...", "info"))
                    for attempt in range(5):
                        result = page.evaluate(js_click_panel(panel_id))
                        if result not in ("not found", "no sap button"):
                            break
                        self._sleep(0.5)
                    self.root.after(0, lambda r=result: self._log(f"Panel: {r}", "info"))

                    # ── 2. Wait for table — poll every 1s, up to 30s ──
                    for _ in range(60):
                        self._sleep(0.3)
                        found = page.evaluate(JS_WAIT_FOR_TABLE)
                        if found == "found":
                            break
                    else:
                        self.root.after(0, lambda: self._log("Table never appeared — skipping", "warning"))
                        self._navigate_back(page)
                        panel_open = False
                        continue

                    panel_open = True

                # ── 3 & 4. Filter then download: Approved → Warehouse → All ──
                filter_ok = self._apply_filters(page, search_term)

                if filter_ok == "NO_COMBO":
                    # Panel has no recognizable status combo — can't verify Approved, skip
                    self.root.after(0, lambda: self._log("No status combo — cannot filter by Approved, skipping panel", "warning"))
                    filter_ok = False

                if filter_ok == "TRY_WAREHOUSE":
                    # No Approved option — try warehouse keeper status
                    filter_ok = self._try_warehouse_filter(page, search_term)

                if not filter_ok:
                    self.root.after(0, lambda: self._log("No data found after all filter attempts — skipping", "warning"))
                    continue

                dl, sk, tot = self._download_rows(page, download_dir, dest_folder)
                grand_downloaded += dl
                grand_skipped += sk
                grand_total += tot

                # ── 4b. Wait for page to settle before next month ──
                is_last_term = (term_idx == len(search_terms) - 1)
                if not is_last_term:
                    self.root.after(0, lambda: self._log("Waiting for page to settle before next month...", "info"))
                    self._sleep(3)

                # ── 5. Navigate back only when switching declarations ──
                is_last_term = (term_idx == len(search_terms) - 1)
                is_last_decl = (decl_idx == len(selected_decls) - 1)
                if is_last_term and not is_last_decl:
                    # switching to a different declaration — go back to main page
                    self._navigate_back(page)
                    panel_open = False
                elif is_last_term and is_last_decl:
                    pass  # all done — leave page as-is

        # ── 6. Combine all files ──
        if grand_downloaded > 0:
            self.root.after(0, lambda: self._log("Combining all files...", "accent"))
            self._combine_files(download_dir)

        _d, _s, _t = grand_downloaded, grand_skipped, grand_total
        self.root.after(0, lambda: self._log(
            f"ALL DONE! {_d} downloaded, {_s} skipped, {_t} total", "success"))
        self.root.after(0, lambda: self._update_stats(progress=100))
        self.root.after(0, lambda: self._scrape_done(_d, _s, _t))

    # ── Sleep helper (checks stop_requested frequently) ───────────────────────

    def _sleep(self, seconds):
        """Sleep in 0.2s chunks so stop_requested is checked frequently."""
        steps = int(seconds / 0.2)
        for _ in range(steps):
            if self.stop_requested:
                return
            time.sleep(0.2)
        remainder = seconds - (steps * 0.2)
        if remainder > 0 and not self.stop_requested:
            time.sleep(remainder)

    def _wait_not_busy(self, page, timeout_s=30):
        """Block until SAP's busy indicator is gone (up to timeout_s seconds)."""
        for _ in range(timeout_s * 2):
            busy = page.evaluate("""
            () => {
                var b = document.querySelector('.sapUiLocalBusyIndicatorAnimation');
                if (b && b.getBoundingClientRect().width > 0) return 'busy';
                return 'idle';
            }
            """)
            if busy == "idle":
                return
            self._sleep(0.5)
        self.root.after(0, lambda: self._log("Busy indicator still present after timeout — continuing anyway", "warning"))

    # ── ApplyFilters ──────────────────────────────────────────────────────────

    def _apply_filters(self, page, search_term):
        search_term = search_term.lower()
        self._sleep(2)

        # ── Status → Approved ──
        status_result = "FAIL"
        for attempt in range(4):
            status_result = page.evaluate(JS_SET_STATUS_APPROVED)
            self.root.after(0, lambda r=status_result, a=attempt: self._log(f"Status attempt {a}: {r}", "info"))
            if status_result == "APPROVED_SET":
                break
            if status_result in ("ARROW_NOT_FOUND", "COMBO_NOT_FOUND"):
                self._sleep(1)
            else:
                break
        self._sleep(1)

        if status_result in ("ARROW_NOT_FOUND", "COMBO_NOT_FOUND"):
            self.root.after(0, lambda r=status_result: self._log(f"Status combo not available ({r}) — skipping", "warning"))
            return "NO_COMBO"

        if status_result == "NO_APPROVED":
            self.root.after(0, lambda: self._log("No Approved option — trying warehouse", "info"))
            return "TRY_WAREHOUSE"

        # ── Search ──
        sv = page.evaluate(js_search(search_term))
        self.root.after(0, lambda v=sv: self._log(f"Search set: {v}", "info"))
        self._sleep(1)

        # ── Page size → 1000 ──
        for attempt in range(4):
            pv = page.evaluate(JS_SET_PAGE_1000)
            self.root.after(0, lambda v=pv, a=attempt: self._log(f"Page size attempt {a}: {v}", "info"))
            if pv == "1000":
                break
            self._sleep(1)

        # ── Click Go ──
        go_result = page.evaluate(JS_CLICK_GO)
        self.root.after(0, lambda r=go_result: self._log(f"Go button: {r}", "info"))
        self._sleep(1)

        # ── Poll every 0.5s up to 30s ──
        check = "NO_DATA"
        for _ in range(60):
            self._sleep(0.5)
            check = page.evaluate(JS_CHECK_NO_DATA)
            if check in ("HAS_DATA", "NO_RECORDS"):
                break

        self.root.after(0, lambda c=check: self._log(f"Data check: {c}", "info"))

        if check == "HAS_DATA":
            self.root.after(0, lambda: self._log("Approved filter — data found", "success"))
            self._sleep(1)  # let row count settle before download starts
            return True

        return False

    def _try_warehouse_filter(self, page, search_term):
        search_term = search_term.lower()
        self.root.after(0, lambda: self._log("Trying warehouse keeper status...", "info"))
        wh = "FAIL"
        for attempt in range(4):
            wh = page.evaluate(JS_SET_STATUS_WAREHOUSE)
            self.root.after(0, lambda r=wh, a=attempt: self._log(f"Warehouse status attempt {a}: {r}", "info"))
            if wh == "WAREHOUSE_SET":
                break
            self._sleep(1)
        if wh != "WAREHOUSE_SET":
            self.root.after(0, lambda: self._log("Warehouse status not available — no data", "warning"))
            return False
        self._sleep(1)
        page.evaluate(js_search(search_term))
        self._sleep(1)
        page.evaluate(JS_SET_PAGE_1000)
        self._sleep(1)
        go_result = page.evaluate(JS_CLICK_GO)
        self.root.after(0, lambda r=go_result: self._log(f"Warehouse Go: {r}", "info"))
        self._sleep(1)

        check = "NO_DATA"
        for _ in range(60):
            self._sleep(0.5)
            check = page.evaluate(JS_CHECK_NO_DATA)
            if check in ("HAS_DATA", "NO_RECORDS"):
                break

        self.root.after(0, lambda c=check: self._log(f"Warehouse data check: {c}", "info"))
        if check == "HAS_DATA":
            self.root.after(0, lambda: self._log("Warehouse filter — data found", "success"))
            return True
        self.root.after(0, lambda: self._log("No data found after all filter attempts", "warning"))
        return False

    # ── Navigate back to Excise Tax main page ─────────────────────────────────

    def _navigate_back(self, page):
        page.evaluate(JS_NAVIGATE_BACK)
        self._sleep(1.5)

    # ── Download all rows (mirrors PAD Downloader) ────────────────────────────

    def _download_rows(self, page, download_dir, dest_folder):
        # Give the portal 3s to start rendering after the filter is applied
        self._sleep(3)

        # Wait for SAP to finish loading before counting rows
        self._wait_not_busy(page)

        # Poll until the SAP table is visible and has rows — up to 20s
        table_id = "TABLE_NOT_FOUND"
        total_rows = 0
        for attempt in range(40):
            table_id = page.evaluate(JS_FIND_TABLE)
            if table_id != "TABLE_NOT_FOUND":
                rc_text = page.evaluate(JS_GET_ROW_COUNT)
                total_rows = int(rc_text) if rc_text.isdigit() else 0
                if total_rows > 0:
                    break
            self.root.after(0, lambda a=attempt: self._log(f"Waiting for table... (attempt {a+1})", "info") if a % 4 == 0 else None)
            self._sleep(0.5)

        if table_id == "TABLE_NOT_FOUND":
            self.root.after(0, lambda: self._log("Table not found after 20s — skipping", "error"))
            return 0, 0, 0

        self.root.after(0, lambda tr=total_rows: self._log(f"Rows to download: {tr}", "success"))

        if total_rows == 0:
            return 0, 0, 0

        # Page size
        ps_text = page.evaluate(JS_GET_PAGE_SIZE)
        page_size = int(ps_text) if ps_text.isdigit() else 100

        self.root.after(0, lambda tr=total_rows: self._update_stats(total=tr))

        row_index = 0
        page_row_index = 0
        downloaded = 0
        skipped = 0
        expected_files = set()
        safe_to_txn = {}

        while row_index < total_rows:
            if self.stop_requested:
                break

            # Pagination
            if page_row_index >= page_size:
                self.root.after(0, lambda: self._log("Next page...", "accent"))
                nr = page.evaluate(JS_CLICK_NEXT)
                if nr in ("NEXT_NOT_FOUND", "NEXT_DISABLED"):
                    break
                self._sleep(0.5)
                page_row_index = 0
                page.evaluate(JS_FIND_TABLE)

            # Scroll
            page.evaluate(js_scroll_to_row(page_row_index))

            # Extract TXN
            txn = page.evaluate(js_extract_txn(page_row_index))
            if txn in ("END", "TABLE_NOT_FOUND", "TABLE_NOT_VISIBLE", "COLUMN_NOT_FOUND", "EMPTY"):
                skipped += 1
                row_index += 1
                page_row_index += 1
                continue

            # Clear popups
            page.evaluate(JS_CLEAR_POPUPS)

            # Click More
            more = page.evaluate(js_click_more(page_row_index))
            if more == "MORE_NOT_FOUND":
                skipped += 1
                _s, _ri, _tn = skipped, row_index, txn
                self.root.after(0, lambda ri=_ri, tn=_tn: self._log(f"Row {ri}: {tn} — no More btn", "warning"))
                row_index += 1
                page_row_index += 1
                continue
            # Export to Excel — intercept download directly via Playwright
            safe = txn.replace("/", "-").replace("\\", "-").replace(":", "-").strip()
            os.makedirs(dest_folder, exist_ok=True)
            dest = os.path.join(dest_folder, f"{safe}.xlsx")
            expected_files.add(f"{safe}.xlsx")
            safe_to_txn[f"{safe}.xlsx"] = txn

            new_file = None
            export_ok = False
            try:
                with page.expect_download(timeout=30000) as dl_info:
                    # Try real Playwright click first
                    clicked = False
                    try:
                        btn = page.locator("bdi:text('Export to Excel')").first
                        if btn.is_visible(timeout=500):
                            btn.click(timeout=1000)
                            clicked = True
                    except Exception:
                        pass
                    if not clicked:
                        page.evaluate(JS_CLICK_EXPORT)
                download = dl_info.value
                # Save directly to destination with correct name
                if not os.path.exists(dest):
                    download.save_as(dest)
                else:
                    alt = dest.replace(".xlsx", "_dup.xlsx")
                    download.save_as(alt)
                    self.root.after(0, lambda t=txn: self._log(f"Duplicate TXN {t} — saved as _dup", "warning"))
                new_file = dest
                export_ok = True
            except Exception as e:
                # Fallback: watch Downloads folder
                self.root.after(0, lambda err=str(e): self._log(f"Download intercept failed: {err} — watching folder", "warning"))
                before_files = set(_list_downloads(download_dir))
                for attempt in range(6):
                    jr = page.evaluate(JS_CLICK_EXPORT)
                    if jr not in ("TEXT_NOT_FOUND", "BUTTON_NOT_FOUND"):
                        export_ok = True
                        break
                    self._sleep(0.3)
                if export_ok:
                    for _ in range(40):
                        self._sleep(0.5)
                        after_files = set(_list_downloads(download_dir))
                        new_files = {f for f in (after_files - before_files)
                                     if not f.endswith(".crdownload") and not f.endswith(".tmp")}
                        if new_files:
                            new_file = max(new_files, key=os.path.getmtime)
                            _wait_for_stable_file(new_file)
                            if not os.path.exists(dest):
                                shutil.move(new_file, dest)
                            else:
                                alt = dest.replace(".xlsx", "_dup.xlsx")
                                shutil.move(new_file, alt)
                                self.root.after(0, lambda t=txn: self._log(f"Duplicate TXN {t} — saved as _dup", "warning"))
                            new_file = dest
                            break

            if new_file is not None:
                downloaded += 1
                _ri, _tn = row_index, txn
                self.root.after(0, lambda ri=_ri, tn=_tn: self._log(f"✓ {tn}", "success"))
            else:
                skipped += 1
                _tn = txn
                self.root.after(0, lambda tn=_tn: self._log(f"Download failed: {tn}", "error"))

            row_index += 1
            page_row_index += 1

            # Update stats
            _d, _s = downloaded, skipped
            self.root.after(0, lambda d=_d, s=_s: self._update_stats(downloaded=d, skipped=s))

        self.root.after(0, lambda d=downloaded: self._log(f"Section done: {d} downloaded", "success"))

        # ── Verify + auto-retry any missing files ──
        if expected_files:
            actual_files = {f for f in os.listdir(dest_folder) if f.endswith(".xlsx")} if os.path.isdir(dest_folder) else set()
            missing = expected_files - actual_files

            if missing:
                self.root.after(0, lambda m=len(missing): self._log(
                    f"{m} file(s) missing — retrying...", "warning"))

                ri = 0
                pri = 0
                while ri < total_rows and missing and not self.stop_requested:
                    if pri >= page_size and total_rows > 1000:
                        nr = page.evaluate(JS_CLICK_NEXT)
                        if nr in ("NEXT_NOT_FOUND", "NEXT_DISABLED"):
                            break
                        self._sleep(0.5)
                        pri = 0
                        page.evaluate(JS_FIND_TABLE)

                    page.evaluate(js_scroll_to_row(pri))
                    txn_r = page.evaluate(js_extract_txn(pri))
                    if txn_r in ("END", "TABLE_NOT_FOUND", "TABLE_NOT_VISIBLE", "COLUMN_NOT_FOUND", "EMPTY"):
                        ri += 1
                        pri += 1
                        continue

                    safe_r = txn_r.replace("/", "-").replace("\\", "-").replace(":", "-").strip()
                    fname_r = f"{safe_r}.xlsx"

                    if fname_r not in missing:
                        ri += 1
                        pri += 1
                        continue

                    self.root.after(0, lambda tn=txn_r: self._log(f"Retrying: {tn}", "accent"))
                    page.evaluate(JS_CLEAR_POPUPS)
                    more_r = page.evaluate(js_click_more(pri))
                    if more_r == "MORE_NOT_FOUND":
                        ri += 1
                        pri += 1
                        continue
                    self._sleep(0.3)

                    dest_r = os.path.join(dest_folder, fname_r)
                    retry_ok = False
                    try:
                        with page.expect_download(timeout=30000) as dl_info:
                            clicked = False
                            try:
                                btn = page.locator("bdi:text('Export to Excel')").first
                                if btn.is_visible(timeout=500):
                                    btn.click(timeout=1000)
                                    clicked = True
                            except Exception:
                                pass
                            if not clicked:
                                page.evaluate(JS_CLICK_EXPORT)
                        dl_r = dl_info.value
                        if not os.path.exists(dest_r):
                            dl_r.save_as(dest_r)
                        else:
                            dl_r.save_as(dest_r + ".tmp")
                            os.remove(dest_r + ".tmp")
                        retry_ok = True
                    except Exception as e:
                        self.root.after(0, lambda err=str(e): self._log(f"Retry failed: {err}", "error"))

                    if retry_ok:
                        missing.discard(fname_r)
                        downloaded += 1
                        self.root.after(0, lambda tn=txn_r: self._log(f"✓ Retry OK: {tn}", "success"))

                    ri += 1
                    pri += 1

            actual_files = {f for f in os.listdir(dest_folder) if f.endswith(".xlsx")} if os.path.isdir(dest_folder) else set()
            still_missing = expected_files - actual_files
            if still_missing:
                self.root.after(0, lambda m=len(still_missing), t=len(expected_files): self._log(
                    f"STILL MISSING {m}/{t} after retry — manual check needed", "error"))
                for mf in sorted(still_missing):
                    self.root.after(0, lambda f=mf, st=safe_to_txn: self._log(
                        f"  FAILED: {st.get(f, f)}", "error"))
            else:
                self.root.after(0, lambda t=len(expected_files): self._log(
                    f"Verified: all {t} files present", "success"))

        return downloaded, skipped, total_rows

    # ── Combine Files (pure Python + openpyxl — no Excel/VBScript needed) ───────

    def _combine_files(self, root_dir):
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        combined_path = os.path.join(root_dir, "ExciseTax_Combined.xlsx")
        self.root.after(0, lambda: self._log("Combining files...", "accent"))

        def _clean_value(val):
            """Convert numeric strings to numbers, preserve leading-zero strings."""
            if not isinstance(val, str):
                return val
            if any(c.isalpha() for c in val):
                return val
            cleaned = "".join(c for c in val if c.isdigit() or c in ".,-")
            cleaned = cleaned.replace(",", "")
            if not cleaned:
                return val
            try:
                if val.lstrip().startswith("0") and len(val.strip()) > 1 and "." not in val:
                    return cleaned  # preserve as string but strip non-numeric chars (matches VBA)
                return float(cleaned) if "." in cleaned else int(cleaned)
            except ValueError:
                return val

        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "CombinedData"

        header_written = False
        next_row = 1
        total_rows_written = 0

        for decl_key, (_, _, folder_name) in PANEL_MAP.items():
            decl_dir = os.path.join(root_dir, folder_name)
            if not os.path.isdir(decl_dir):
                continue
            # Walk all subfolders (month subfolders like "january 2025")
            xlsx_files = []
            for dirpath, _, filenames in os.walk(decl_dir):
                for f in sorted(filenames):
                    if f.endswith(".xlsx") and not f.startswith("~$"):
                        xlsx_files.append(os.path.join(dirpath, f))
            for fpath in xlsx_files:
                fname = os.path.basename(fpath)
                fn_clean = os.path.splitext(fname)[0]
                month_label = os.path.basename(os.path.dirname(fpath)).title()
                try:
                    wb_src = load_workbook(fpath, data_only=True)
                except Exception as e:
                    self.root.after(0, lambda p=fname, err=str(e): self._log(f"Skipping {p}: {err}", "warning"))
                    continue

                for sheet in wb_src.sheetnames:
                    ws_src = wb_src[sheet]
                    rows = list(ws_src.iter_rows(values_only=True))
                    if not rows:
                        continue
                    src_headers = list(rows[0])
                    data_rows = rows[1:]

                    if not header_written:
                        header = ["DeclarationType", "Month", "FileName", "SheetName"] + src_headers
                        ws_out.append(header)
                        next_row = 2
                        header_written = True

                    for row in data_rows:
                        cleaned = [_clean_value(v) for v in row]
                        ws_out.append([folder_name, month_label, fn_clean, sheet] + cleaned)
                        next_row += 1
                        total_rows_written += 1

                wb_src.close()

        if not header_written:
            self.root.after(0, lambda: self._log("No files to combine", "warning"))
            return

        # ── Formatting ──
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="8B1A2B")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws_out[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ws_out.freeze_panes = "A2"

        # Auto-width based on content
        col_widths = {}
        for row in ws_out.iter_rows():
            for cell in row:
                if cell.value is not None:
                    col = cell.column
                    col_widths[col] = max(col_widths.get(col, 0), min(len(str(cell.value)), 40))
        for col, width in col_widths.items():
            ws_out.column_dimensions[get_column_letter(col)].width = width + 2

        try:
            wb_out.save(combined_path)
            self.root.after(0, lambda p=combined_path, n=total_rows_written: self._log(
                f"Combined {n} rows → {p}", "success"))
            self.root.after(0, lambda p=combined_path: messagebox.showinfo(
                "Complete", f"All files combined into:\n{p}"))
        except Exception as e:
            self.root.after(0, lambda err=str(e): self._log(f"Save failed: {err}", "error"))

    # ── Cleanup ───────────────────────────────────────────────────────────────

    def _on_close(self):
        self.stop_requested = True
        self._pw_queue.put(None)  # shut down worker thread
        try:
            if self.pw_browser:
                self.pw_browser.close()
            if self.pw_instance:
                self.pw_instance.stop()
            if self._chrome_proc:
                self._chrome_proc.terminate()
        except Exception:
            pass
        self.root.destroy()


if __name__ == "__main__":
    try:
        ExciseScraperApp()
    except Exception as e:
        import traceback
        print("\n" + "="*60)
        print("ERROR — the app crashed before it could open:")
        print("="*60)
        traceback.print_exc()
        print("="*60)
        print("\nMost likely fix: run these commands in Command Prompt:")
        print("  pip install playwright openpyxl")
        input("\nPress Enter to close...")
